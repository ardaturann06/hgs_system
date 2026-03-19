from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from typing import List, Optional
from pydantic import BaseModel
import io
import openpyxl

from database import get_db, init_db, Vehicle, Passage, User
from schemas import (
    VehicleCreate, VehicleUpdate, VehicleOut,
    PassageCreate, PassageOut, VehicleReport
)
from auth import (
    hash_password, verify_password, create_token,
    get_current_user, require_admin
)

app = FastAPI(title="HGS Filo Takip Sistemi", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()
    # İlk admin kullanıcıyı oluştur
    db = next(get_db())
    if not db.query(User).filter(User.username == "admin").first():
        admin = User(
            username="admin",
            full_name="Sistem Yöneticisi",
            hashed_password=hash_password("admin123"),
            role="admin",
            is_active=True
        )
        db.add(admin)
        db.commit()
        print("✓ Admin kullanıcı oluşturuldu → admin / admin123")
    db.close()


# ──────────────────────────────────────────
# AUTH
# ──────────────────────────────────────────

class UserCreate(BaseModel):
    username: str
    full_name: Optional[str] = None
    password: str
    role: str = "user"

class UserOut(BaseModel):
    id: int
    username: str
    full_name: Optional[str]
    role: str
    is_active: bool
    created_at: datetime
    class Config:
        from_attributes = True

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class TokenOut(BaseModel):
    access_token: str
    token_type: str
    username: str
    full_name: Optional[str]
    role: str


@app.post("/auth/login", response_model=TokenOut, tags=["Auth"])
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form.username).first()
    if not user or not verify_password(form.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Kullanıcı adı veya şifre hatalı")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Hesap devre dışı")
    token = create_token({"sub": user.username})
    return TokenOut(
        access_token=token,
        token_type="bearer",
        username=user.username,
        full_name=user.full_name,
        role=user.role
    )


@app.get("/auth/me", response_model=UserOut, tags=["Auth"])
def me(current_user: User = Depends(get_current_user)):
    return current_user


@app.get("/users", response_model=List[UserOut], tags=["Kullanıcılar"])
def list_users(current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    return db.query(User).all()


@app.post("/users", response_model=UserOut, tags=["Kullanıcılar"])
def create_user(data: UserCreate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == data.username).first():
        raise HTTPException(status_code=400, detail="Bu kullanıcı adı zaten alınmış")
    user = User(
        username=data.username,
        full_name=data.full_name,
        hashed_password=hash_password(data.password),
        role=data.role
    )
    db.add(user); db.commit(); db.refresh(user)
    return user


@app.put("/users/{user_id}", response_model=UserOut, tags=["Kullanıcılar"])
def update_user(user_id: int, data: UserUpdate, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if data.full_name is not None: user.full_name = data.full_name
    if data.password:              user.hashed_password = hash_password(data.password)
    if data.role is not None:      user.role = data.role
    if data.is_active is not None: user.is_active = data.is_active
    db.commit(); db.refresh(user)
    return user


@app.delete("/users/{user_id}", tags=["Kullanıcılar"])
def delete_user(user_id: int, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="Admin silinemez")
    db.delete(user); db.commit()
    return {"message": f"{user.username} silindi"}


# ──────────────────────────────────────────
# EXCEL IMPORT / EXPORT
# ──────────────────────────────────────────

@app.post("/vehicles/import-excel", tags=["Araçlar"])
async def import_vehicles_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx veya .xls dosyası yükleyin")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")

    added, skipped, errors = [], [], []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        plate = str(row[0]).strip().upper().replace(" ", "").replace("-", "")
        if not plate:
            continue

        hgs_tag    = str(row[1]).strip() if len(row) > 1 and row[1] else None
        owner_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
        try:
            balance = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
        except (ValueError, TypeError):
            balance = 0.0

        if db.query(Vehicle).filter(Vehicle.plate == plate).first():
            skipped.append(plate)
            continue

        try:
            vehicle = Vehicle(plate=plate, hgs_tag=hgs_tag, owner_name=owner_name, balance=balance)
            db.add(vehicle)
            db.commit()
            added.append(plate)
        except Exception as e:
            db.rollback()
            errors.append({"plate": plate, "error": str(e)})

    return {
        "added": len(added),
        "skipped": len(skipped),
        "errors": len(errors),
        "added_plates": added,
        "skipped_plates": skipped,
        "error_details": errors
    }


@app.get("/vehicles/export-excel", tags=["Araçlar"])
def export_vehicles_excel(db: Session = Depends(get_db), _=Depends(get_current_user)):
    vehicles = db.query(Vehicle).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Araçlar"

    # Header stili
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="FF6B35")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Plaka", "HGS Tag", "Sürücü / Sahip", "Bakiye (₺)", "Kayıt Tarihi"]
    col_widths = [15, 15, 25, 15, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = w

    ws.row_dimensions[1].height = 22

    for r, v in enumerate(vehicles, start=2):
        row_data = [
            v.plate,
            v.hgs_tag or "",
            v.owner_name or "",
            round(v.balance, 2),
            v.created_at.strftime("%d.%m.%Y %H:%M") if v.created_at else ""
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col in (1,2,4,5) else "left", vertical="center")
            if col == 4:
                cell.number_format = '#,##0.00 "₺"'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=araclar.xlsx"}
    )


@app.get("/vehicles/template-excel", tags=["Araçlar"])
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Şablon"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="FF6B35")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Plaka *", "HGS Tag", "Sürücü / Sahip", "Bakiye (₺)"]
    col_widths = [15, 15, 25, 15]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[chr(64 + col)].width = w

    # Örnek satırlar
    examples = [
        ["34ABC123", "HGS-001", "Ahmet Yılmaz", 200.0],
        ["06XY789",  "HGS-002", "Mehmet Demir", 150.0],
        ["35DEF456", "",        "Ayşe Kaya",    0.0],
    ]
    from openpyxl.styles import PatternFill as PF
    ex_fill = PF("solid", fgColor="FFF3EE")
    for r, row in enumerate(examples, start=2):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = ex_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col in (1,2,4) else "left")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=arac_sablonu.xlsx"}
    )


# ──────────────────────────────────────────
# ARAÇLAR
# ──────────────────────────────────────────

@app.post("/vehicles", response_model=VehicleOut, tags=["Araçlar"])
def create_vehicle(data: VehicleCreate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    plate = data.plate.upper().replace(" ", "")
    if db.query(Vehicle).filter(Vehicle.plate == plate).first():
        raise HTTPException(status_code=400, detail="Bu plaka zaten kayıtlı")
    vehicle = Vehicle(plate=plate, hgs_tag=data.hgs_tag, owner_name=data.owner_name, balance=data.balance or 0.0)
    db.add(vehicle); db.commit(); db.refresh(vehicle)
    return vehicle


@app.get("/vehicles", response_model=List[VehicleOut], tags=["Araçlar"])
def list_vehicles(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Vehicle).all()


@app.get("/vehicles/{plate}", response_model=VehicleOut, tags=["Araçlar"])
def get_vehicle(plate: str, db: Session = Depends(get_db), _=Depends(get_current_user)):
    plate = plate.upper().replace(" ", "")
    vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    return vehicle


@app.put("/vehicles/{plate}", response_model=VehicleOut, tags=["Araçlar"])
def update_vehicle(plate: str, data: VehicleUpdate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    plate = plate.upper().replace(" ", "")
    vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    if data.hgs_tag is not None:    vehicle.hgs_tag = data.hgs_tag
    if data.owner_name is not None: vehicle.owner_name = data.owner_name
    if data.balance is not None:    vehicle.balance = data.balance
    db.commit(); db.refresh(vehicle)
    return vehicle


@app.delete("/vehicles/{plate}", tags=["Araçlar"])
def delete_vehicle(plate: str, db: Session = Depends(get_db), _=Depends(require_admin)):
    plate = plate.upper().replace(" ", "")
    vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    db.delete(vehicle); db.commit()
    return {"message": f"{plate} silindi"}


# ──────────────────────────────────────────
# GEÇİŞLER
# ──────────────────────────────────────────

@app.post("/passages", response_model=PassageOut, tags=["Geçişler"])
def add_passage(data: PassageCreate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    plate = data.plate.upper().replace(" ", "")
    vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    passage = Passage(
        vehicle_id=vehicle.id,
        location=data.location,
        amount=data.amount,
        passed_at=data.passed_at or datetime.utcnow(),
        note=data.note
    )
    vehicle.balance -= data.amount
    db.add(passage); db.commit(); db.refresh(passage)
    return passage


@app.get("/passages", response_model=List[PassageOut], tags=["Geçişler"])
def list_passages(plate: Optional[str] = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    query = db.query(Passage)
    if plate:
        plate = plate.upper().replace(" ", "")
        vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
        if not vehicle:
            raise HTTPException(status_code=404, detail="Araç bulunamadı")
        query = query.filter(Passage.vehicle_id == vehicle.id)
    return query.order_by(Passage.passed_at.desc()).all()


# ──────────────────────────────────────────
# RAPORLAR
# ──────────────────────────────────────────

@app.get("/reports/{plate}", response_model=VehicleReport, tags=["Raporlar"])
def vehicle_report(plate: str, db: Session = Depends(get_db), _=Depends(get_current_user)):
    plate = plate.upper().replace(" ", "")
    vehicle = db.query(Vehicle).filter(Vehicle.plate == plate).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Araç bulunamadı")
    passages = db.query(Passage).filter(Passage.vehicle_id == vehicle.id).order_by(Passage.passed_at.desc()).all()
    return VehicleReport(
        plate=vehicle.plate, owner_name=vehicle.owner_name, hgs_tag=vehicle.hgs_tag,
        balance=vehicle.balance, total_passages=len(passages),
        total_spent=sum(p.amount for p in passages), passages=passages
    )


@app.get("/reports", tags=["Raporlar"])
def fleet_summary(db: Session = Depends(get_db), _=Depends(get_current_user)):
    vehicles = db.query(Vehicle).all()
    result = []
    for v in vehicles:
        total = db.query(func.sum(Passage.amount)).filter(Passage.vehicle_id == v.id).scalar() or 0.0
        count = db.query(func.count(Passage.id)).filter(Passage.vehicle_id == v.id).scalar() or 0
        result.append({
            "plate": v.plate, "owner_name": v.owner_name, "hgs_tag": v.hgs_tag,
            "balance": v.balance, "total_passages": count, "total_spent": round(total, 2)
        })
    return result
